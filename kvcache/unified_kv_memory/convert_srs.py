import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def parse_markdown_row(line):
    # Strip leading/trailing | and split by |
    parts = line.strip().split('|')
    if line.strip().startswith('|'):
        parts = parts[1:]
    if line.strip().endswith('|'):
        parts = parts[:-1]
    return [p.strip() for p in parts]

def is_separator_row(row):
    return all(re.match(r'^:?-+:?$', cell) for cell in row) if row else False

def parse_markdown_tables(md_path):
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines()
        
    tables = []
    current_section = "Sheet1"
    
    i = 0
    while i < len(lines):
        line = lines[i]
        # Track section headers
        if line.strip().startswith('#'):
            header_match = re.match(r'^#+\s*(.*)$', line.strip())
            if header_match:
                current_section = header_match.group(1).strip()
            i += 1
            continue
            
        if line.strip().startswith('|'):
            # Collect all consecutive table rows
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                table_lines.append(lines[i])
                i += 1
            
            parsed_rows = [parse_markdown_row(tl) for tl in table_lines]
            
            if len(parsed_rows) >= 3:
                header = parsed_rows[0]
                if is_separator_row(parsed_rows[1]):
                    data_rows = parsed_rows[2:]
                else:
                    data_rows = parsed_rows[1:]
                
                tables.append({
                    'section': current_section,
                    'header': header,
                    'rows': data_rows
                })
            elif len(parsed_rows) > 0:
                tables.append({
                    'section': current_section,
                    'header': parsed_rows[0],
                    'rows': parsed_rows[1:]
                })
        else:
            i += 1
            
    return tables

def make_valid_sheet_name(name):
    # Remove characters not allowed in Excel sheet names: : \ / ? * [ ]
    cleaned = re.sub(r'[:\\/?*\[\]]', '', name)
    return cleaned[:31].strip()

def main():
    md_path = 'srs.md'
    output_path = 'srs.xlsx'
    
    print(f"Reading tables from {md_path}...")
    tables = parse_markdown_tables(md_path)
    print(f"Found {len(tables)} tables.")
    
    wb = Workbook()
    # Store default active sheet to remove later
    default_sheet = wb.active
    
    font_name = 'Arial'
    
    # Styles Definition
    header_font = Font(name=font_name, size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', start_color='0D2137', end_color='0D2137')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name=font_name, size=10, bold=False, color='000000')
    
    even_fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFFFF')
    odd_fill = PatternFill(fill_type='solid', start_color='FFEEF2FF', end_color='FFEEF2FF') # Light lavender/blue-gray
    
    thin_border_side = Side(border_style='thin', color='FFBBBBBB')
    border_all_thin = Border(
        top=thin_border_side,
        bottom=thin_border_side,
        left=thin_border_side,
        right=thin_border_side
    )
    
    priority_styles = {
        'P0': {
            'font': Font(name=font_name, size=10, bold=True, color='CC0000'),
            'fill': PatternFill(fill_type='solid', start_color='FFFFEBEE', end_color='FFFFEBEE')
        },
        'P1': {
            'font': Font(name=font_name, size=10, bold=True, color='E65100'),
            'fill': PatternFill(fill_type='solid', start_color='FFFBE9E7', end_color='FFFBE9E7')
        },
        'P2': {
            'font': Font(name=font_name, size=10, bold=True, color='1565C0'),
            'fill': PatternFill(fill_type='solid', start_color='FFE3F2FD', end_color='FFE3F2FD')
        },
        'P3': {
            'font': Font(name=font_name, size=10, bold=True, color='555555'),
            'fill': PatternFill(fill_type='solid', start_color='FFF5F5F5', end_color='FFF5F5F5')
        },
        'P4': {
            'font': Font(name=font_name, size=10, bold=True, color='777777'),
            'fill': PatternFill(fill_type='solid', start_color='FAFAFA', end_color='FAFAFA')
        }
    }
    
    for table in tables:
        sheet_title = make_valid_sheet_name(table['section'])
        print(f"Creating sheet: {sheet_title}")
        ws = wb.create_sheet(title=sheet_title)
        
        # 1. Write Header Row
        ws.append(table['header'])
        ws.row_dimensions[1].height = 27
        
        for c_idx, cell_val in enumerate(table['header'], 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border_all_thin
            
        # 2. Write Data Rows
        for r_idx, row_data in enumerate(table['rows'], 2):
            ws.append(row_data)
            
            # Determine base fill (Zebra Striping)
            base_fill = even_fill if (r_idx % 2 == 0) else odd_fill
            
            for c_idx, cell_val in enumerate(row_data, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.fill = base_fill
                cell.border = border_all_thin
                
                # Check alignments
                header_col = table['header'][c_idx - 1]
                h_align = 'left'
                
                # Check if Priority Cell
                if header_col in ['优先级', 'Priority']:
                    h_align = 'center'
                    val_str = str(cell_val).strip()
                    if val_str in priority_styles:
                        cell.font = priority_styles[val_str]['font']
                        cell.fill = priority_styles[val_str]['fill']
                
                # Check other fields for center alignment
                elif any(kw in header_col for kw in ['ID', '标识', '层级', '模块ID', '需求唯一标识', '顶层业务模块 (Top-Level Module)']):
                    h_align = 'center'
                    
                cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=True)
                
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # 3. Auto-fit column widths using smart sizing
        for c_idx in range(1, len(table['header']) + 1):
            header_col = table['header'][c_idx - 1]
            col_letter = get_column_letter(c_idx)
            
            col_vals = [str(ws.cell(row=r, column=c_idx).value or '') for r in range(2, len(table['rows']) + 2)]
            max_len = max([len(v) for v in col_vals] + [len(header_col)])
            
            # Smart width allocation based on keywords
            if any(kw in header_col for kw in ['ID', '标识', '需求唯一标识']):
                width = 25
            elif any(kw in header_col for kw in ['优先级', 'Priority']):
                width = 12
            elif any(kw in header_col for kw in ['层级', '模块说明', '核心模块', '顶层业务模块 (Top-Level Module)']):
                width = 20
            elif any(kw in header_col for kw in ['描述', '要点', '职责', '模块核心职责描述 (第一性原理定位)']):
                width = 55
            elif any(kw in header_col for kw in ['场景', '痛点', '指标', '验收方法']):
                width = 30
            else:
                width = min(max(max_len * 1.2, 12), 50)
                
            ws.column_dimensions[col_letter].width = width
            
    # Remove default sheet if we added our own
    if len(wb.sheetnames) > 1 and 'Sheet' in wb.sheetnames:
        wb.remove(default_sheet)
        
    wb.save(output_path)
    print(f"Successfully saved to {output_path}")

if __name__ == '__main__':
    main()
