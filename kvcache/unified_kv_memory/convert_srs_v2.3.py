import re
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def parse_markdown_row(line):
    # Strip leading/trailing | and split by |
    parts = line.strip().split('|')
    if line.strip().startswith('|'):
        parts = parts[1:]
    if line.strip().endswith('|'):
        parts = parts[:-1]
    return [p.strip() for p in parts]

def parse_markdown_row_by_comma(line):
    parts = line.strip().split(',')
    return [p.strip() for p in parts]

def is_separator_row(row):
    return all(re.match(r'^:?-+:?$', cell) for cell in row) if row else False

def parse_markdown_tables(md_path):
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines()
        
    tables = []
    current_section = "Sheet1"
    
    i = 0
    while i < len(lines):
        line = lines[i]
        # Track section headers
        if line.strip().startswith('#'):
            header_match = re.match(r'^#+\s*(.*)$', line.strip())
            if header_match:
                current_section = header_match.group(1).strip()
            i += 1
            continue
            
        if line.strip().startswith('|'):
            # Collect all consecutive table rows
            table_lines = []
            while i < len(lines) and lines[i].strip().startswith('|'):
                table_lines.append(lines[i])
                i += 1
            
            parsed_rows = [parse_markdown_row(tl) for tl in table_lines]
            
            if len(parsed_rows) >= 3:
                header = parsed_rows[0]
                if is_separator_row(parsed_rows[1]):
                    data_rows = parsed_rows[2:]
                else:
                    data_rows = parsed_rows[1:]
                
                tables.append({
                    'section': current_section,
                    'header': header,
                    'rows': data_rows
                })
            elif len(parsed_rows) > 0:
                tables.append({
                    'section': current_section,
                    'header': parsed_rows[0],
                    'rows': parsed_rows[1:]
                })
        else:
            i += 1
            
    return tables

def parse_new_requirements(md_path):
    if not os.path.exists(md_path):
        print(f"Warning: File not found: {md_path}")
        return []
        
    with open(md_path, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines()
        
    rows = []
    for line in lines:
        if line.strip().startswith('|'):
            row = parse_markdown_row(line)
            if not is_separator_row(row) and not any(kw in (row[0] if row else '') for kw in ['需求唯一标识', '唯一标识']):
                if any(cell.strip() for cell in row):
                    rows.append(row)
    return rows

def parse_reviewed_requirements(txt_path):
    if not os.path.exists(txt_path):
        print(f"Warning: File not found: {txt_path}")
        return []
        
    with open(txt_path, 'r', encoding='utf-8') as f:
        lines = f.read().splitlines()
        
    rows = []
    for line in lines:
        if ',' in line and not line.startswith('需求唯一标识') and not line.startswith('##'):
            row = parse_markdown_row_by_comma(line)
            if any(cell.strip() for cell in row):
                rows.append(row)
    return rows

def make_valid_sheet_name(name):
    # Remove characters not allowed in Excel sheet names: : \ / ? * [ ]
    cleaned = re.sub(r'[:\\/?*\[\]]', '', name)
    return cleaned[:31].strip()

def clean_module_name(name):
    # Remove circled number prefix like ⑥
    return re.sub(r'^[①②③④⑤⑥⑦⑧⑨⑩]\s*', '', name).strip()

def main():
    md_path = 'srs.md'
    reviewed_path = '新增评审后的需求.txt'
    output_path = 'srs_v2.3.xlsx'
    
    print(f"Reading base tables from {md_path}...")
    tables = parse_markdown_tables(md_path)
    print(f"Found {len(tables)} base tables.")
    
    # 1. Merge original new requirements (v2.2)
    additions = {
        'L2 KVConnector层': ['新增需求L2.md'],
        'L3 传输管理层': ['新增需求L3A.md', '新增需求L3B.md', '新增需求L3C.md', '新增需求L3D.md', '新增需求L3E.md'],
        'L4 底层传输层': ['新增需求L4.md']
    }
    
    for table in tables:
        section = table['section']
        if section in additions:
            print(f"\nMerging v2.2 additions for sheet: {section}...")
            for file_name in additions[section]:
                new_rows = parse_new_requirements(file_name)
                table['rows'].extend(new_rows)
                print(f"  Parsed {len(new_rows)} rows from {file_name}")
            
    # 2. Merge reviewed requirements (v2.3)
    print(f"\nReading reviewed requirements from {reviewed_path}...")
    reviewed_rows = parse_reviewed_requirements(reviewed_path)
    print(f"Parsed {len(reviewed_rows)} reviewed rows in total.")
    
    for table in tables:
        section = table['section']
        section_to_prefix = {
            'L2 KVConnector层': 'L2-',
            'L3 传输管理层': 'L3-',
            'L4 底层传输层': 'L4-'
        }
        if section in section_to_prefix:
            prefix = section_to_prefix[section]
            matching_reviewed = [r for r in reviewed_rows if r[0].startswith(prefix)]
            table['rows'].extend(matching_reviewed)
            print(f"  Merged {len(matching_reviewed)} reviewed rows into {section}")
            print(f"  Total rows for {section} is now {len(table['rows'])}")
            
    # New requirements mapping to top-level business modules (TM1 to TM6)
    new_id_to_tm = {
        # L2 (from v2.2)
        'L2-KV-PublishCommit-033': 'TM1',
        'L2-KV-AttachHandle-034': 'TM5',
        'L2-KV-ConsumeEligibility-035': 'TM2',
        'L2-KV-SemanticIdentity-036': 'TM1',
        'L2-KV-StateErrorCode-037': 'TM6',
        'L2-KV-PartialAttachPlan-038': 'TM4',
        
        # L3 (from v2.2)
        'L3-MS-KVObjectStateMachine-058': 'TM3',
        'L3-MS-ReplicaPlacementState-059': 'TM2',
        'L3-MS-ConsumeEligibility-060': 'TM2',
        'L3-SE-PlacementResolver-061': 'TM1',
        'L3-MS-MigrationInterlock-062': 'TM3',
        'L3-CO-AttachDetachLease-063': 'TM5',
        'L3-CO-VisibilityReadyBitmap-064': 'TM5',
        'L3-CO-PublishCommitLog-065': 'TM5',
        'L3-MS-InvalidationTombstone-066': 'TM2',
        'L3-MS-StaleHitGuard-067': 'TM2',
        'L3-CO-ReplicaQuarantine-068': 'TM5',
        'L3-MS-PrefixDirectorySchema-069': 'TM2',
        'L3-MS-HotLocalIndex-070': 'TM2',
        'L3-MS-RangeBatchLookup-071': 'TM2',
        'L3-SE-QueryPlanFastPath-072': 'TM1',
        'L3-MS-DirectoryShardPolicy-073': 'TM2',
        'L3-MS-MetadataCompactLayout-074': 'TM2',
        'L3-MC-KVBlockStorageSchema-075': 'TM3',
        'L3-MC-ExtentManifest-076': 'TM3',
        'L3-MC-PageExtentAllocator-077': 'TM3',
        'L3-MC-LayoutTransformPlan-078': 'TM3',
        'L3-SE-DescriptorFromManifest-079': 'TM1',
        'L3-CO-KVOwnershipGC-080': 'TM5',
        'L3-MS-StateAwarePrefetch-081': 'TM4',
        'L3-MS-StateAwareEviction-082': 'TM3',
        'L3-OB-KVStateTrace-083': 'TM6',
        'L3-CTRL-KVInspectAPI-084': 'TM6',
        
        # L4 (from v2.2)
        'L4-CO-AtomicRemapPrimitive-065': 'TM5',
        'L4-CO-ExtentVisibilityFence-066': 'TM5',
        'L4-HW-StorageLayoutCapability-067': 'TM1',
        'L4-QO-StateAwareTrafficClass-068': 'TM5',
        'L4-FT-ReplicaIntegrityCheck-069': 'TM6',
        'L4-MC-PersistentExtentHandle-070': 'TM4',
        'L4-RDMA-RemoteExtentHandle-071': 'TM4',
        'L4-UB-DirectViewGuard-072': 'TM3',
        'L4-CO-DrainAndQuiesce-073': 'TM5',
        'L4-SC-SecureExtentRelease-074': 'TM5',
        
        # L2 (from v2.3)
        'L2-CONN-CostAwareReturn-039': 'TM1',
        'L2-CONN-BufferContract-040': 'TM1',
        
        # L3 (from v2.3)
        'L3-CO-RefCountLifecycle-086': 'TM5',
        'L3-MS-AtomicPublishVisibility-087': 'TM2',
        'L3-MS-TTFTIndexLayout-088': 'TM2',
        'L3-SE-MultiReplicaResolver-089': 'TM1',
        'L3-CO-MigrationRCULock-090': 'TM5',
        
        # L4 (from v2.3)
        'L4-HW-NICSpecConstraint-075': 'TM1'
    }
    
    # Supplement descriptions for the Core Module Dictionary
    dict_updates = {
        ('L2', '统一协议与标准接口'): '统一 KVConnector Protocol 与标准数据结构，支持三阶段发布与语义兼容校验，支持拥塞链路自适应返回与算力执行内存解耦。',
        ('L2', '共享访问与故障控制'): '多消费者共享、View Lease、fallback 契约、路径追踪，支持标准 attach 句柄与统一状态错误码。',
        ('L2', '本地元数据缓存与快速判定'): '本地元数据缓存、两级前缀索引、批量前缀查询与可消费状态判定。',
        ('L2', '传输描述与布局协商'): 'KV 布局协商、批量传输 descriptor 与 coalescing，支持部分前缀命中复用。',
        ('L3', '统一存储池与分层管理'): '集群统一内存池、分层管理、淘汰与分配，支持对象状态机、迁移互锁、块存储结构与布局转换。',
        ('L3', '前缀目录与元数据平面'): '分布式前缀索引与元数据平面，支持副本状态、可消费判定、最终一致可见性发布与 TTFT 极速检索。',
        ('L3', '语义策略与路径管理'): '统一传输语义与策略引擎，支持最优位置解析、查询快路径、传输描述符自动生成与多副本最优路径解析。',
        ('L3', '一致性、隔离与QoS'): '一致发布、多租户隔离与 QoS，支持租约、可见性控制、发布日志、副本隔离、GC 回收、销毁驱逐保护与无锁 RCU 迁移。',
        ('L3', '传输编排与广播'): '传输编排与广播，支持 UBLink 路由、一对多分发、热点复制与状态感知预取。',
        ('L3', '全路径观测与故障追踪'): '全路径性能观测与故障追踪，支持状态轨迹追踪与统一运维排障接口。',
        ('L4', '一致性、安全与可靠性'): '一致性、安全与可靠性，支持原子重映射、可见性 fence、状态感知流量隔离、损坏检测与安全释放。',
        ('L4', 'Fabric能力与统一路由'): '硬件能力表、地址转换与混合 Fabric 路由，支持布局能力描述与底层网络软硬件协同 QP 规格约束。',
        ('L4', 'DPU卸载与分层存储I/O'): 'DPU 传输卸载与分层存储 I/O，支持持久化 Extent 句柄寻址。',
        ('L4', 'RDMA与零拷贝传输'): 'RDMA与零拷贝传输，支持远端 Extent 句柄安全寻址与注册内存池。',
        ('L4', '统一内存与内存语义访问'): '统一内存与内存语义访问，支持 CPU/NPU 统一内存池、直访保护与 memory view 消费。'
    }

    wb = Workbook()
    default_sheet = wb.active
    
    font_name = 'Arial'
    
    # Styles Definition
    header_font = Font(name=font_name, size=10, bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', start_color='0D2137', end_color='0D2137')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name=font_name, size=10, bold=False, color='000000')
    
    even_fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFFFF')
    odd_fill = PatternFill(fill_type='solid', start_color='FFEEF2FF', end_color='FFEEF2FF')
    
    thin_border_side = Side(border_style='thin', color='FFBBBBBB')
    border_all_thin = Border(
        top=thin_border_side,
        bottom=thin_border_side,
        left=thin_border_side,
        right=thin_border_side
    )
    
    priority_styles = {
        'P0': {
            'font': Font(name=font_name, size=10, bold=True, color='CC0000'),
            'fill': PatternFill(fill_type='solid', start_color='FFFFEBEE', end_color='FFFFEBEE')
        },
        'P1': {
            'font': Font(name=font_name, size=10, bold=True, color='E65100'),
            'fill': PatternFill(fill_type='solid', start_color='FFFBE9E7', end_color='FFFBE9E7')
        },
        'P2': {
            'font': Font(name=font_name, size=10, bold=True, color='1565C0'),
            'fill': PatternFill(fill_type='solid', start_color='FFE3F2FD', end_color='FFE3F2FD')
        },
        'P3': {
            'font': Font(name=font_name, size=10, bold=True, color='555555'),
            'fill': PatternFill(fill_type='solid', start_color='FFF5F5F5', end_color='FFF5F5F5')
        },
        'P4': {
            'font': Font(name=font_name, size=10, bold=True, color='777777'),
            'fill': PatternFill(fill_type='solid', start_color='FAFAFA', end_color='FAFAFA')
        }
    }
    
    print("\nWriting to Excel sheets...")
    for table in tables:
        section = table['section']
        sheet_title = make_valid_sheet_name(section)
        print(f"Creating sheet: {sheet_title}")
        ws = wb.create_sheet(title=sheet_title)
        
        # 1. Write Header Row
        ws.append(table['header'])
        ws.row_dimensions[1].height = 27
        
        for c_idx, cell_val in enumerate(table['header'], 1):
            cell = ws.cell(row=1, column=c_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border_all_thin
            
        # 2. Write Data Rows
        for r_idx, row_data in enumerate(table['rows'], 2):
            modified_row = list(row_data)
            
            # Apply supplements to 横向模块分类 and 01_核心模块字典
            if section == '横向模块分类':
                tm_cell = modified_row[0]
                tm_match = re.match(r'^(TM\d+)', tm_cell)
                if tm_match:
                    tm_id = tm_match.group(1)
                    # Col 3: index 2 (L1), Col 4: index 3 (L2), Col 5: index 4 (L3), Col 6: index 5 (L4)
                    for col_idx, layer in enumerate(['L1', 'L2', 'L3', 'L4'], 2):
                        new_ids = [req_id for req_id, mapped_tm in new_id_to_tm.items() 
                                   if mapped_tm == tm_id and req_id.startswith(layer + '-')]
                        if new_ids:
                            existing_val = modified_row[col_idx].strip()
                            if existing_val == '-' or not existing_val:
                                modified_row[col_idx] = ', '.join(new_ids)
                            else:
                                existing_ids = [i.strip() for i in existing_val.split(',') if i.strip()]
                                for nid in new_ids:
                                    if nid not in existing_ids:
                                        existing_ids.append(nid)
                                modified_row[col_idx] = ', '.join(existing_ids)
            elif section == '01_核心模块字典':
                layer = modified_row[0]
                module = modified_row[1]
                clean_module = clean_module_name(module)
                key = (layer, clean_module)
                if key in dict_updates:
                    modified_row[2] = dict_updates[key]
            
            ws.append(modified_row)
            
            # Determine base fill (Zebra Striping)
            base_fill = even_fill if (r_idx % 2 == 0) else odd_fill
            
            for c_idx, cell_val in enumerate(modified_row, 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = data_font
                cell.fill = base_fill
                cell.border = border_all_thin
                
                # Check alignments
                header_col = table['header'][min(c_idx - 1, len(table['header']) - 1)]
                h_align = 'left'
                
                # Check if Priority Cell
                if header_col in ['优先级', 'Priority']:
                    h_align = 'center'
                    val_str = str(cell_val).strip()
                    if val_str in priority_styles:
                        cell.font = priority_styles[val_str]['font']
                        cell.fill = priority_styles[val_str]['fill']
                
                # Check other fields for center alignment
                elif any(kw in header_col for kw in ['ID', '标识', '层级', '模块ID', '需求唯一标识', '顶层业务模块 (Top-Level Module)']):
                    h_align = 'center'
                    
                cell.alignment = Alignment(horizontal=h_align, vertical='center', wrap_text=True)
                
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
        
        # 3. Auto-fit column widths using smart sizing
        for c_idx in range(1, len(table['header']) + 1):
            header_col = table['header'][c_idx - 1]
            col_letter = get_column_letter(c_idx)
            
            col_vals = [str(ws.cell(row=r, column=c_idx).value or '') for r in range(2, len(table['rows']) + 2)]
            max_len = max([len(v) for v in col_vals] + [len(header_col)])
            
            # Smart width allocation based on keywords
            if any(kw in header_col for kw in ['ID', '标识', '需求唯一标识']):
                width = 25
            elif any(kw in header_col for kw in ['优先级', 'Priority']):
                width = 12
            elif any(kw in header_col for kw in ['层级', '模块说明', '核心模块', '顶层业务模块 (Top-Level Module)']):
                width = 20
            elif any(kw in header_col for kw in ['描述', '要点', '职责', '模块核心职责描述 (第一性原理定位)']):
                width = 55
            elif any(kw in header_col for kw in ['场景', '痛点', '指标', '验收方法', 'L推理调度层需求 ID', 'LKVConnector 层需求 ID', 'L传输管理层需求 ID', 'L底层传输层需求 ID']):
                width = 40
            else:
                width = min(max(max_len * 1.2, 12), 50)
                
            ws.column_dimensions[col_letter].width = width
            
    # Remove default sheet if we added our own
    if len(wb.sheetnames) > 1 and 'Sheet' in wb.sheetnames:
        wb.remove(default_sheet)
        
    wb.save(output_path)
    print(f"\nSuccessfully merged and saved to {output_path}")

if __name__ == '__main__':
    main()
